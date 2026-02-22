#!/usr/bin/env python3
"""
MIDAS Excel Schema Loader and Work Order Generator
"""

import argparse
import random
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import load_workbook, Workbook

# ===================================================================
# FILE / SHEET CONFIGURATION
# ===================================================================
SCHEMA_FILE = r"MIDAS_Excel_Schema.xlsx"  # <- updated to your actual file
SHEET_INSTALLATIONS = "Installations"
SHEET_FACILITIES = "Facilities"
SHEET_ROOMS = "Rooms"
SHEET_REQUESTING_ORGS = "RequestingOrgs"
SHEET_TRADES = "Trades"

# ===================================================================
# UTILITIES
# ===================================================================
def read_excel_sheet(file_path: str, sheet_name: str):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    data = []
    headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def pick_weighted(options: list, weights: list):
    return random.choices(options, weights=weights, k=1)[0]

def fill_template(template: str, template_values: dict):
    result = template
    for key, (lo, hi) in template_values.items():
        placeholder = "{" + key + "}"
        while placeholder in result:
            result = result.replace(placeholder, str(random.randint(lo, hi)), 1)
    return result

# ===================================================================
# LOAD SCHEMA DATA
# ===================================================================
installations_data = read_excel_sheet(SCHEMA_FILE, SHEET_INSTALLATIONS)
facilities_data = read_excel_sheet(SCHEMA_FILE, SHEET_FACILITIES)
rooms_data = read_excel_sheet(SCHEMA_FILE, SHEET_ROOMS)
requesting_orgs_data = read_excel_sheet(SCHEMA_FILE, SHEET_REQUESTING_ORGS)
trades_data = read_excel_sheet(SCHEMA_FILE, SHEET_TRADES)

# Convert to simple lists/dicts
INSTALLATIONS = {row['InstallationName']: row for row in installations_data}
REQUESTING_ORGS = [row['OrgName'] for row in requesting_orgs_data]
TRADES = [row['TradeName'] for row in trades_data]

ORG_WEIGHTS = [row.get('Weight', 1) for row in requesting_orgs_data]
TRADE_WEIGHTS = [row.get('Weight', 1) for row in trades_data]

TEMPLATE_VALUES = {
    "temp": (78, 95), "setpoint": (68, 72), "ahu": (1, 8), "ahu2": (3, 12),
    "unit": (1, 6), "zone": (1, 12), "panel": (1, 16), "imbalance": (5, 15),
    "voltage": (2, 8), "psi": (15, 45), "months": (14, 24), "cam": (1, 32),
    "post": (1, 4), "pct": (10, 25), "pole": (1, 20),
}

# ===================================================================
# WORK ORDER GENERATOR (simplified for demo)
# ===================================================================
def generate_work_order(seq_num: int, base_date: datetime):
    installation = pick_weighted(list(INSTALLATIONS.keys()), [0.5]*len(INSTALLATIONS))
    facility = random.choice(facilities_data)
    room = random.choice(rooms_data)['RoomName']
    trade = pick_weighted(TRADES, TRADE_WEIGHTS)
    org = pick_weighted(REQUESTING_ORGS, ORG_WEIGHTS)
    request_dt = base_date + timedelta(hours=random.randint(0, 72))
    return {
        "Work Order #": f"WO-{seq_num:04d}",
        "Installation": installation,
        "Facility": facility['FacilityName'],
        "Room": room,
        "Trade": trade,
        "Requesting Org": org,
        "Request DateTime": request_dt.strftime("%Y-%m-%d %H:%M")
    }

# ===================================================================
# MAIN
# ===================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--count", type=int, default=10)
    parser.add_argument("--output", type=str, default="my_work_orders.xlsx")
    args = parser.parse_args()

    work_orders = [generate_work_order(i+1, datetime.now()) for i in range(args.count)]

    wb = Workbook()
    ws = wb.active
    ws.title = "WorkOrders"

    headers = list(work_orders[0].keys())
    ws.append(headers)
    for wo in work_orders:
        ws.append([wo[h] for h in headers])

    wb.save(args.output)
    print(f"{len(work_orders)} work orders generated and saved to {args.output}")